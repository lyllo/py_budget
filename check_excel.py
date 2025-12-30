import sys
import os
from openpyxl import load_workbook

sys.path.append(os.path.abspath("src"))
import load.files as files

def check_excel():
    path = "out/final.xlsx"
    if not os.path.exists(path):
        print("File not found.")
        return
    
    wb = load_workbook(path)
    ws = wb["Summary"]
    print(f"Total rows in Summary: {ws.max_row}")
    
    # Print last 5 rows
    print("\nLast 5 rows:")
    for row in ws.iter_rows(min_row=max(1, ws.max_row - 5), values_only=True):
        print(row)

if __name__ == "__main__":
    check_excel()
