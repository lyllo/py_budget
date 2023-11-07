from openpyxl import load_workbook
from fuzzywuzzy import fuzz
import openai

# 
# BUSCA SIMPLES (ANÁLISE LÉXICA)
#

# Carrega as colunas do Excel referentes a ITEM e CATEGORIA
def carrega_dicionario():

    # Carrega o arquivo Excel BUDGET_SET23 com transações até o mês passado
    workbook = load_workbook('history.xlsx')

    # Seleciona a planilha Summary desejada
    worksheet = workbook['Summary']

    # Cria um dicionário vazio
    dicionario = {}

    # Percorre as células das colunas B e D iniciando pela linha 2, onde B = Item e C = Categoria.
    # Exemplo ["iFood" = "Alimentação"]
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        chave = row[1]
        valor = row[3]
        dicionario[chave] = valor
    
    # Fecha o arquivo
    workbook.close()

    # Retorna o dicionário
    return dicionario

#
# BUSCA POR SIMILARIDADE
#

def busca_palavras_parecidas(palavra, lista_palavras, limite):
    palavras_parecidas = []
    for palavra_lista in lista_palavras:
        similaridade = fuzz.ratio(palavra, palavra_lista)
        if similaridade >= limite:
            palavras_parecidas.append(palavra_lista)
    return palavras_parecidas

#
# BUSCA POR AI (SW 3.0)
#

# Definir chave de API do OpenAI
openai.api_key = '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'

# Função para fazer a chamada à API e obter a resposta da LLM
def interagir_com_llm(prompt):
    response = openai.Completion.create(
        engine='text-davinci-003',
        prompt=prompt,
        max_tokens=100,
        temperature=0.7,
        n=1,
        stop=None,
        timeout=10
    )
    return response.choices[0].text.strip()

# Função para carregar as categorias como string no prompt da LLM
def carrega_categorias():
    categorias = ["ÁGUA",
                "ALIMENTAÇÃO",
                "ALUGUEL",
                "BEBÊ",
                "CARRO",
                "CASA",
                # "COMPRAS",
                "DOMÉSTICA",
                "GÁS",
                "INTERNET",
                "LUZ",
                "MERCADO",
                # "OUTROS",
                "PET",
                "SAÚDE",
                "SERVIÇOS",
                "STREAMING",
                "TRANSPORTE",
                "VIAGENS"]
    return(", ".join(categorias))

# Preparar o prompt para um registro
def prepara_prompt(estabelecimentos):
    prompt = "Retorne uma lista de strings em Python contendo, dentre as seguintes categorias \"" + carrega_categorias() + "\", quais seriam as mais adequadas para classificar as compras realizadas com um cartão de crédito, cujos estabelecimentos comerciais se chamam, respectivamente " + str(estabelecimentos) + ", preenchendo com uma string vazia quando a entrada também for uma string vazia."
    # print("Prompt: " + prompt)
    # print(str(estabelecimentos))
    return prompt

def preenche_categorias_com_respostas_da_ai(lista_de_registros, resposta):
    # print(type(resposta))
    for indice, registro in enumerate(lista_de_registros):
        if "categoria" not in registro:
            registro['categoria'] = resposta[indice]

# Fecha a lista com "]" caso a LLM não retorne com esse caractere
def verificar_resposta(resposta):
    if resposta[-2] != "]":
        return resposta[:-1] + "]" + resposta[-1:]

def busca_categoria_com_ai(lista_de_registros):
    # Preparar a lista de itens separada por vírgula
    lista_de_registros_para_ai = []
    for registro in lista_de_registros:
        if "categoria" not in registro:
            lista_de_registros_para_ai.append(registro['item'])
        else:
            lista_de_registros_para_ai.append('')
    # Monta o prompt
    prompt = prepara_prompt(lista_de_registros_para_ai)
    # Chama a LLM
    resposta = interagir_com_llm(prompt)
    resposta_limpa = verificar_resposta(resposta)
    # Imprime a resposta (alterar para preencher as categorias)
    # print("Response: " + resposta.upper())
   
    # Percorrer a lista de categorias para preencher com os resultados
    preenche_categorias_com_respostas_da_ai(lista_de_registros, eval(resposta_limpa.upper()))

# prompt = prepara_prompt("[\'AIRBNB\', \'PÃO DE AÇÚCAR\', \'\', \'\', \'UBER\']")