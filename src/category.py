from openpyxl import load_workbook
from fuzzywuzzy import fuzz
import pickle
import configparser
import os
import ai
from datetime import datetime

# Configura os paths dos arquivos que serão utilizados
ROOT_DIR = os.path.dirname(
    os.path.dirname(
        os.path.abspath(__file__)
    )
)
PATH_TO_CONFIG_FILE = os.path.join(ROOT_DIR, 'config.ini')
PATH_TO_HISTORY_FILE = os.path.join(ROOT_DIR, 'data\\history.xlsx')
PATH_TO_BIN_FILE = os.path.join(ROOT_DIR, 'data\\dados.bin')

# limite de similaridade, valor de 75 encontrado por inspeção manual
limite = 75

# Le as feature toggles do arquivo de configuração
config = configparser.ConfigParser()
config.read(PATH_TO_CONFIG_FILE)
load_xlsx = config['default']['load_xlsx']
simple_match = config['default']['simple_match']
similar_match = config['default']['similar_match']
ai_match = config['default']['ai_match']
verbose = config['default']['verbose']

# 
# BUSCA SIMPLES (ANÁLISE LÉXICA)
#

from load.db import conecta_bd

# ... (imports)

# Carrega as colunas do Excel referentes a ITEM e CATEGORIA e mescla com DB
def carrega_dicionario():

    # Cria um dicionário vazio
    dicionario = {}

    if(load_xlsx == "true"):

        # Carrega o arquivo Excel history com transações até o mês passado
        workbook = load_workbook(PATH_TO_HISTORY_FILE)

        # Seleciona a planilha Summary desejada
        worksheet = workbook['Summary']

        # Percorre as células das colunas B e H iniciando pela linha 2, onde B = Item e H = Categoria.
        # Exemplo ["iFood" = "Alimentação"]
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            chave = row[1]
            valor = row[7]
            if chave and valor:
                dicionario[chave] = valor
        
        # Fecha o arquivo
        workbook.close()

    else:
        # Carrega os dados salvos em um arquivo binário
        if os.path.exists(PATH_TO_BIN_FILE):
            with open(PATH_TO_BIN_FILE, 'rb') as arquivo:
                dicionario = pickle.load(arquivo)

    # [OPTIMIZATION] Carrega também do Banco de Dados (MariaDB) para garantir que o aprendizado recente seja considerado
    try:
        conn = conecta_bd()
        cursor = conn.cursor()
        cursor.execute("SELECT item, categoria FROM transactions WHERE categoria IS NOT NULL AND categoria != '' AND categoria != 'OUTROS'")
        rows = cursor.fetchall()
        
        count_db_learn = 0
        for row in rows:
            item_db = row[0]
            cat_db = row[1]
            # Prioridade para o DB (mais recente/confiável) ou apenas preenche lacunas?
            # Vamos preencher lacunas e sobrescrever, pois DB é a fonte da verdade final.
            if item_db:
                dicionario[item_db] = cat_db
                count_db_learn += 1
        
        conn.close()
        if verbose == "true":
            print(f"[Smart Learning] Carregados {count_db_learn} itens categorizados diretamente do DB.")

    except Exception as e:
        if verbose == "error":
            print(f"Erro ao carregar dicionário do DB: {e}")

    # Atualiza o binário com o consolidado (Excel + DB)
    try:
        with open(PATH_TO_BIN_FILE, 'wb') as arquivo:
            pickle.dump(dicionario, arquivo)
    except Exception as e:
        pass

    # Retorna o dicionário
    return dicionario

#
# BUSCA POR SIMILARIDADE
#

def busca_palavras_parecidas(palavra, lista_palavras, limite):
    palavras_parecidas = []
    for palavra_lista in lista_palavras:
        similaridade = fuzz.ratio(palavra, palavra_lista)
        if similaridade >= limite:
            palavras_parecidas.append(palavra_lista)
    return palavras_parecidas
    
# Retira pontos e deixa a palavra em maiúsculo
def limpa_resposta(resposta):
    # Remove o caractere da string
    resposta_limpa = resposta.replace('.', '')

    # Torna toda a string em maiúsculas
    resposta_limpa = resposta_limpa.upper()

    return resposta_limpa 

def busca_categoria_com_ai(lista_de_registros):
    total_input_tokens = 0
    total_output_tokens = 0
    calls_made = 0

    for registro in lista_de_registros:
        # Só chama a IA se a categoria ainda estiver vazia ou for "Outros"
        if not registro.get('categoria') or registro.get('categoria').upper() == 'OUTROS':
            
            registro_para_ai = registro['item']
            prompt = ai.prepara_prompt(registro_para_ai)
            resposta, usage = ai.interagir_com_llm(prompt)
            
            if usage:
                total_input_tokens += usage.prompt_token_count
                total_output_tokens += usage.candidates_token_count
                calls_made += 1

            resposta_limpa = limpa_resposta(resposta)
            
            # Validação: só aceita se estiver na lista oficial (opcional, mas bom para consistência)
            if resposta_limpa in [c.upper() for c in ai.OFFICIAL_CATEGORIES]:
                registro['categoria'] = resposta_limpa
                registro['categoria_fonte'] = "ai_gpt"
                if(verbose == "true"):
                    print(f"[IA] Categorizado: {registro_para_ai} -> {resposta_limpa}")
            else:
                if(verbose == "true"):
                    print(f"[IA] Categoria inválida retornada: {resposta_limpa} para {registro_para_ai}")
    
    if calls_made > 0:
        # Gemini Flash pricing (USD): $0.075 per 1M input tokens, $0.30 per 1M output tokens
        cost_input = (total_input_tokens / 1_000_000) * 0.075
        cost_output = (total_output_tokens / 1_000_000) * 0.30
        total_cost = cost_input + cost_output
        print(f"\n[AI Usage Summary - Gemini]")
        print(f"Total Calls: {calls_made}")
        print(f"Input Tokens: {total_input_tokens} (${cost_input:,.6f})")
        print(f"Output Tokens: {total_output_tokens} (${cost_output:,.6f})")
        print(f"Estimated Cost: ${total_cost:,.6f}\n")

def fill(lista_de_registros):

    # Busca por categorias
    if (simple_match == "true"):

        # Imprime timestamp do início do carregamento do dicionário de categorias
        if(verbose == "debug"):
            print("[" + datetime.now().strftime("%H:%M:%S") +"] Iniciando carregamento do dicionário...")

        # Carrega dicionário de categorias
        lista_de_categorias = carrega_dicionario()

        # Imprime timestamp do término do carregamento do dicionário
        if(verbose == "debug"):
            print("[" + datetime.now().strftime("%H:%M:%S") +"] Dicionário carregado em memória.")

        num_simple_matches = 0
        num_similar_matches = 0

        for registro in lista_de_registros:

            registro['categoria_fonte'] = ''
            
            # Faz a busca exata
            if registro['item'] in lista_de_categorias:
                registro['categoria'] = lista_de_categorias[registro['item']]
                registro['categoria_fonte'] = 'history_exact'
                num_simple_matches += 1
            
            # Faz a busca por similaridade
            else:
                if(similar_match == "true"):
                    
                    palavras_parecidas = busca_palavras_parecidas(registro['item'], lista_de_categorias.keys(), limite)
                    
                    # só estamos interessados quando a busca encontra apenas 1 elemento e não mais de 1
                    if len(palavras_parecidas) == 1:
                        
                        registro['categoria'] = lista_de_categorias[palavras_parecidas[0]]
                        registro['categoria_fonte'] = 'history_similar'
                        num_similar_matches += 1

        if (simple_match == "debug"):
            if(verbose == "true"):
                print(str(num_simple_matches) + " simple matches em " + str(len(lista_de_registros)) + " registros, eficiência de " + str(round(100*(num_simple_matches/len(lista_de_registros)),2)) + "%.")

        if (similar_match == "debug"):
            if(verbose == "true"):
                print(str(num_similar_matches) + " similar matches quando limite = " + str(limite) + " em " + str(len(lista_de_registros)) + " registros, eficiência de " + str(round(100*(num_similar_matches/len(lista_de_registros)),2)) + "%.")

    # Faz a busca com ai

    if (ai_match == "true"):
        busca_categoria_com_ai(lista_de_registros)

def aprende_categoria(item, categoria):
    """Atualiza o dicionário de categorias com um novo aprendizado manual e persiste no binário."""
    if not item or not categoria:
        return
    
    try:
        # Tenta carregar o dicionário atual (preferencialmente do binário)
        # Se o arquivo não existir, inicia um novo
        if os.path.exists(PATH_TO_BIN_FILE):
            with open(PATH_TO_BIN_FILE, 'rb') as arquivo:
                dicionario = pickle.load(arquivo)
        else:
            dicionario = {}
        
        # Só atualiza e salva se houver mudança real
        if dicionario.get(item) != categoria:
            dicionario[item] = categoria
            with open(PATH_TO_BIN_FILE, 'wb') as arquivo:
                pickle.dump(dicionario, arquivo)
            
            if verbose == "true":
                print(f"[Aprendizado] Item '{item}' vinculado à categoria '{categoria}' em dados.bin")
                
    except Exception as e:
        if verbose == "error":
            print(f"Erro ao salvar aprendizado: {e}")