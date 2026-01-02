from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from datetime import datetime
from datetime import date
import xlrd
import load.db as db
import os, platform
import configparser

# Caminho do arquivo atual
current_file_path = os.path.abspath(__file__)

# Caminho da raiz do projeto
ROOT_DIR = os.path.abspath(os.path.join(current_file_path, "../../.."))
PATH_TO_FINAL_OUTPUT_FILE = os.path.join(ROOT_DIR, 'out\\final.xlsx')

# Caminho para arquivo de configuração
PATH_TO_CONFIG_FILE = os.path.join(ROOT_DIR, 'config.ini')

# Lê as feature toggles do arquivo de configuração
config = configparser.ConfigParser()
config.read(PATH_TO_CONFIG_FILE)
verbose = config['default']['verbose']

# Inclui linhas em arquivo Excel
def incluir_linhas_em_excel(nome_arquivo, nome_planilha, linhas):

    new_file = True

    try:
        # Tenta carregar o arquivo existente
        wb = load_workbook(nome_arquivo)
        new_file = False

    except FileNotFoundError:
        # Se o arquivo não existe, cria um novo
        wb = Workbook()

        # Se o arquivo não existe, vamos precisar de um cabeçalho
        linhas.insert(0, ['DATA', 'ITEM', 'DETALHE', 'OCORRÊNCIA', 'VALOR', 'CARTÃO', 'PARCELA', 'CATEGORIA', 'TAG', 'MEIO'])

    try:
        ws = wb[nome_planilha]

    except KeyError:
        # Cria a sheet 'Summary'
        ws = wb.create_sheet(nome_planilha)

    # Inicializa o contador de linhas da planilha
    if new_file == True:
        num_linha_excel = 1
    else:
        num_linha_excel = ws.max_row + 1

    # Inclui as linhas no arquivo do Excel
    for linha in linhas:
        ws.append(linha)
        
        # Pinta a fonte das linhas que contêm transações no futuro de cinza
        if isinstance(linha[0], date):
            data_linha = linha[0]
            data_hoje = datetime.now().date()
            if data_linha > data_hoje:
                for cell in ws[num_linha_excel]:
                    cell.font = Font(color="808080")

        # Formata como moeda em BRL a coluna E (5) do VALOR
        for cell in ws[num_linha_excel]:
            if cell.column == 5:
                cell.number_format = 'R$ #,##0.00'

        # Pintar de vermelho a fonte das células das categorias preenchidas por AI
        if 'ai_gpt' in linha:
            for cell in ws[num_linha_excel]:
                # A coluna F (6) é da CATEGORIA
                if cell.column == 6:
                    # print("Vou pintar a célula da linha " + str(num_linha_excel) + " e coluna " + str(cell.column))
                    cell.font = Font(color='FF0000')

        num_linha_excel += 1

    formata_planilha(ws)

    # Salva os stats
    # save_stats(wb)

    # Salva o arquivo do Excel
    wb.save(nome_arquivo)

# Substituir linhas em arquivo Excel (chamado pelo método dump)
def substituir_linhas_em_excel(nome_arquivo, nome_planilha, linhas):

    new_file = True

    # O try foi pra carregar o arquivo, não tratava incialmente ter o arquivo, mas sem a sheet 'Summary'
    try:
        # Tenta carregar o arquivo existente
        wb = load_workbook(nome_arquivo)

        # Instancia a sheet 'Summary'
        ws = wb[nome_planilha]

        # Remove a sheet 'Summary'
        wb.remove(ws)

        new_file = False

    # Não existe o arquivo
    except FileNotFoundError:
        # Se o arquivo não existe, cria um novo
        wb = Workbook()

    # Existe o arquivo mas não a sheet
    except KeyError:
        # Cria a sheet 'Summary'
        ws = wb.create_sheet(nome_planilha)

    ws = wb.create_sheet(nome_planilha)

    # Criando o cabeçalho
    linhas.insert(0, ['DATA', 'ITEM', 'DETALHE', 'OCORRÊNCIA', 'VALOR', 'CARTÃO', 'PARCELA', 'CATEGORIA', 'TAG', 'MEIO'])

    # Batch style objects
    font_gray = Font(color="808080")
    font_red = Font(color='FF0000')
    data_hoje = datetime.now().date()
    currency_format = 'R$ #,##0.00'

    # Inicializa o contador de linhas da planilha
    num_linha_excel = 1

    # Inclui as linhas no arquivo do Excel
    for linha in linhas:
        ws.append(linha)
        
        # Access the just appended row for formatting
        current_row = ws[num_linha_excel]

        # 1. Row-level Formatting (Single Pass - ONLY when needed)
        # Currency Formatting (Column E / index 4) - Mandatory
        current_row[4].number_format = currency_format

        # Future Coloring (Slow but necessary for specific rows)
        if isinstance(linha[0], date) and linha[0] > data_hoje:
            for cell in current_row:
                cell.font = font_gray
        
        # AI Category Coloring (Category is Column H / index 7)
        elif 'ai_gpt' in linha:
            current_row[7].font = font_red

        num_linha_excel += 1

    formata_planilha(ws)

    if verbose == "true":
        # Exclui a linha de cabeçalho
        print(f"\tRegistros salvos: {ws.max_row-1}")

    # Salva o arquivo do Excel
    wb.save(nome_arquivo)

# Formata a planilha
def formata_planilha(sheet):
    
    # 1. Column Dimensions
    col_widths = {
        'A': 20, 'B': 40, 'C': 40, 'D': 20, 'E': 20, 
        'F': 20, 'G': 20, 'H': 20, 'I': 20, 'J': 20
    }
    for col, width in col_widths.items():
        sheet.column_dimensions[col].width = width

    # 2. Global Row Height (Fast Scale)
    sheet.sheet_format.defaultRowHeight = 15

    # 3. Header Formatting (First row)
    # Re-apply header styles specifically to row 1
    header_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_align = Alignment(horizontal='center', vertical='center')

    for col in range(1, 11):
        cell = sheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    # 4. Features
    sheet.auto_filter.ref = sheet.dimensions
    sheet.freeze_panes = sheet['A2']

# Carregar arquivo de texto
def ler_arquivo(nome_arquivo):
    linhas = []
    with open(nome_arquivo, 'r', -1, 'utf-8') as arquivo:
        for linha in arquivo:
            linhas.append(linha.strip())
    return linhas

# Carregar arquivo xlsx
def ler_arquivo_xlsx(nome_arquivo, nome_planilha):
    # Carregar o arquivo Excel
    workbook = load_workbook(nome_arquivo)

    # Selecionar a planilha ativa
    sheet = workbook.get_sheet_by_name(nome_planilha)

    # Iterar sobre as linhas da planilha
    return (row for row in sheet.iter_rows(values_only=True))

# Carregar arquivo xls
def ler_arquivo_xls(nome_arquivo):
    # Carregar o arquivo Excel
    workbook = xlrd.open_workbook(nome_arquivo)

    # Selecionar a planilha ativa
    sheet = workbook.sheet_by_index(0)

    # Iterar sobre as linhas da planilha
    for row in range(sheet.nrows):
        yield sheet.row_values(row)

# Salvar dados recém carregados em excel (puxando do DB)
def salva_excel(nome_arquivo, nome_planilha):

    transactions = db.fetch_recent_transactions(nome_planilha)

    if verbose == "true":
        print(f"\tRegistros lidos (DB): {len(transactions)}")

    if (len(transactions) > 0):

        # Função de chave para a ordenação
        def chave_de_ordenacao(dic):
            return dic['data']

        # Ordenar a lista de dicionários pela chave 'data'
        lista_ordenada = sorted(transactions, key=chave_de_ordenacao)

        # Transforma a lista de dicionários em uma lista de listas, sem os nomes das chaves
        lista_de_listas = [list(item.values()) for item in lista_ordenada]

        nome_planilha = "Summary"

        # Salva os dados em arquivo excel
        incluir_linhas_em_excel(nome_arquivo, nome_planilha, lista_de_listas)

# Salvar dados recém lidos diretamente em excel (sem passar pelo DB)
def salva_lista_excel(nome_arquivo, nome_planilha, lista_de_registros):

    if verbose == "true":
        print(f"\tSalvando {len(lista_de_registros)} registros diretamente em {os.path.basename(nome_arquivo)}...")

    if (len(lista_de_registros) > 0):

        # Função de chave para a ordenação
        def chave_de_ordenacao(dic):
            # Normalizar data para garantir que seja date ou datetime comparável
            d = dic['data']
            if isinstance(d, datetime):
                return d.date()
            return d

        # Ordenar a lista de dicionários pela chave 'data'
        lista_ordenada = sorted(lista_de_registros, key=chave_de_ordenacao)

        # Transforma a lista de dicionários em uma lista de listas
        # Ordem: ['DATA', 'ITEM', 'DETALHE', 'OCORRÊNCIA', 'VALOR', 'CARTÃO', 'PARCELA', 'CATEGORIA', 'TAG', 'MEIO']
        lista_de_listas = []
        for reg in lista_ordenada:
            lista_de_listas.append([
                reg.get('data'),
                reg.get('item'),
                reg.get('detalhe', ''),
                reg.get('ocorrencia_dia', 1),
                reg.get('valor'),
                reg.get('cartao', ''),
                reg.get('parcela', ''),
                reg.get('categoria', ''),
                reg.get('tag', ''),
                reg.get('meio', nome_planilha) # Usa o meio passado ou o do registro
            ])

        nome_planilha_final = "Summary"

        # Se for o arquivo final, usamos incluir (append) ou substituir? 
        # Geralmente final.xlsx é o consolidado. Se estamos bypassando DB, 
        # talvez o usuário queira concatenar.
        if "final.xlsx" in nome_arquivo.lower():
            incluir_linhas_em_excel(nome_arquivo, nome_planilha_final, lista_de_listas)
        else:
            substituir_linhas_em_excel(nome_arquivo, nome_planilha_final, lista_de_listas)

def get_modification_time(file_path):
    if platform.system() == 'Windows':
        modification_time = os.path.getmtime(file_path)
        # Tenta atualizar o mtime no DB, mas ignora se falhar (pois estamos bypassando)
        try:
            db.update_mtime(file_path, modification_time)
        except:
            pass
        return modification_time
    else:
        print("Este exemplo funciona apenas no Windows, a obtenção da data de criação pode variar em outros sistemas operacionais.")


def dump_history():

    transactions = db.fetch_transactions()

    print(f"\tRegistros lidos: {len(transactions)}")

    if (len(transactions) > 0):

        # Função de chave para a ordenação
        def chave_de_ordenacao(dic):
            return dic['data']

        # Ordenar a lista de dicionários pela chave 'data'
        lista_ordenada = sorted(transactions, key=chave_de_ordenacao)

        # Transforma a lista de dicionários em uma lista de listas, sem os nomes das chaves
        lista_de_listas = [list(item.values()) for item in lista_ordenada]

        nome_planilha = "Summary"

        # Salva os dados em arquivo excel
        substituir_linhas_em_excel(PATH_TO_FINAL_OUTPUT_FILE, nome_planilha, lista_de_listas)

        if verbose == "true":
            timestamp = datetime.now().strftime("%H:%M:%S")
            print(f"\n[{timestamp}] Fim do 'dump' do DB em XLSX.")

def save_stats(wb):

    stats = db.fetch_stats()

    nome_planilha = 'Stats'

    # Verificar se a planilha já existe para apagar e criá-la se não existir
    if nome_planilha in wb.sheetnames:
        sheet_to_remove = wb[nome_planilha]
        wb.remove_sheet(sheet_to_remove)

    sheet = wb.create_sheet(title=nome_planilha)

    # Transforma a lista de dicionários em uma lista de listas, sem os nomes das chaves
    linhas_stats = [list(item.values()) for item in stats]

    # Insere o cabeçalho
    linhas_stats.insert(0, ['MEIO', 'ÚLTIMA TRANSAÇÃO', 'ÚLTIMA ATUALIZAÇÃO', 'ÚLTIMO ARQUIVO', 'TOTAL TRANSAÇÕES'])

    for linha in linhas_stats:
        sheet.append(linha)