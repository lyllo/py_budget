from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl import load_workbook
import xlrd
from datetime import datetime
from datetime import date
import db

# Inclui linhas em arquivo Excel
def incluir_linhas_em_excel(nome_arquivo, nome_planilha, linhas):

    try:
        # Tenta carregar o arquivo existente
        wb = load_workbook(nome_arquivo)

    except FileNotFoundError:
        # Se o arquivo não existe, cria um novo
        wb = Workbook()
        # Adiciona uma planilha padrão (você pode remover ou modificar conforme necessário)
        wb.create_sheet(title=nome_planilha)
    
    # Verificar se a planilha já existe ou criá-la se não existir
    if nome_planilha in wb.sheetnames:
        sheet = wb[nome_planilha]
    else:
        sheet = wb.create_sheet(title=nome_planilha)

    # Inicializa o contador de linhas da planilha
    num_linha_excel = 1

    # Inclui as linhas no arquivo do Excel
    for linha in linhas:
        sheet.append(linha)
        
        # Pintar a fonte das linhas que contêm transações no futuro de cinza
        if isinstance(linha[0], date):
            data_linha = linha[0]
            data_hoje = datetime.now().date()
            if data_linha > data_hoje:
                for cell in sheet[num_linha_excel]:
                    cell.font = Font(color="808080")

        # Formata como moeda em BRL a coluna E (5) do VALOR
        for cell in sheet[num_linha_excel]:
            if cell.column == 5:
                cell.number_format = 'R$ #,##0.00'

        # Pintar de vermelho a fonte das células das categorias preenchidas por AI
        if 'ai_gpt' in linha:
            for cell in sheet[num_linha_excel]:
                # A coluna F (6) é da CATEGORIA
                if cell.column == 6:
                    # print("Vou pintar a célula da linha " + str(num_linha_excel) + " e coluna " + str(cell.column))
                    cell.font = Font(color='FF0000')

        num_linha_excel += 1

    # Salva o arquivo do Excel
    wb.save(nome_arquivo)

# Carregar arquivo de texto
def ler_arquivo(nome_arquivo):
    linhas = []
    with open(nome_arquivo, 'r', -1, 'utf-8') as arquivo:
        for linha in arquivo:
            linhas.append(linha.strip())
    return linhas

# Carregar arquivo xlsx
def ler_arquivo_xlsx(nome_arquivo, nome_planilha):
    # Carregar o arquivo Excel
    workbook = load_workbook(nome_arquivo)

    # Selecionar a planilha ativa
    sheet = workbook.get_sheet_by_name(nome_planilha)

    # Iterar sobre as linhas da planilha
    return (row for row in sheet.iter_rows(values_only=True))

# Carregar arquivo xls
def ler_arquivo_xls(nome_arquivo):
    # Carregar o arquivo Excel
    workbook = xlrd.open_workbook(nome_arquivo)

    # Selecionar a planilha ativa
    sheet = workbook.sheet_by_index(0)

    # Iterar sobre as linhas da planilha
    for row in range(sheet.nrows):
        yield sheet.row_values(row)

# Salvar dados recém carregados em excel temporário

def salva_excel_temporario(nome_arquivo, nome_planilha, timestamp):

    transactions = db.fetch_transactions(timestamp)

    # Função de chave para a ordenação
    def chave_de_ordenacao(dic):
        return dic['data']

    # Ordenar a lista de dicionários pela chave 'data'
    lista_ordenada = sorted(transactions, key=chave_de_ordenacao)

    # Transforma a lista de dicionários em uma lista de listas, sem os nomes das chaves
    lista_de_listas = [list(item.values()) for item in lista_ordenada]

    # Adiciona o cabeçalho à lista de listas
    lista_de_listas.insert(0, ['DATA', 'ITEM', 'DETALHE', 'OCORRÊNCIA', 'VALOR', 'CARTÃO', 'PARCELA', 'CATEGORIA', 'TAG'])

    # Salva os dados em arquivo excel
    incluir_linhas_em_excel(nome_arquivo, nome_planilha, lista_de_listas)